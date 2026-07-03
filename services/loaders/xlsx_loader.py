from pathlib import Path
import openpyxl


def extract_xlsx(file_path: Path):
    """
    Extract text from Excel workbook.
    """

    workbook = openpyxl.load_workbook(
        file_path,
        data_only=True
    )

    sheets = []

    for sheet in workbook.worksheets:

        sheets.append(f"=== Sheet: {sheet.title} ===")

        for row in sheet.iter_rows(values_only=True):

            values = [
                str(cell)
                for cell in row
                if cell is not None
            ]

            if values:
                sheets.append(" | ".join(values))

    return {
        "filename": file_path.name,
        "document_type": "Excel",
        "pages": len(workbook.sheetnames),
        "text": "\n".join(sheets)
    }


def extract_all_xlsx(folder: Path):

    documents = []

    for file in folder.glob("*.xlsx"):
        documents.append(extract_xlsx(file))

    return documents